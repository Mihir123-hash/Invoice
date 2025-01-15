from flask import Flask, jsonify, render_template, request, redirect, url_for, flash, make_response
import pandas as pd
from io import BytesIO
from flask import send_file
import mysql.connector
from datetime import date
import os
import string
from flask_sqlalchemy import SQLAlchemy
import random
from dotenv import load_dotenv
from flask_mail import Mail, Message
from functools import wraps
from xhtml2pdf import pisa
from num2words import num2words
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import numbers

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('APP_SECRET_KEY')  # Secret key for flash messages

db_config = {
    'host': os.getenv('DB_HOST'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'database': os.getenv('DB_NAME')
}

# Configure Flask-Mail
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER')
app.config['MAIL_PORT'] = os.getenv('MAIL_PORT') # Convert to integer
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS') == 'True'  # Convert string to boolean
app.config['MAIL_USE_SSL'] = os.getenv('MAIL_USE_SSL') == 'True'  # Convert string to boolean
mail = Mail(app)

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('SQLALCHEMY_DATABASE_URI')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = os.getenv('SQLALCHEMY_TRACK_MODIFICATIONS') == 'False'  # Convert string to boolean
db = SQLAlchemy(app)

# Function to get a MySQL database connection
def get_db_connection():
    conn = mysql.connector.connect(**db_config)
    return conn

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

# User model
class Users(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)  # Added user_name field
    otp = db.Column(db.String(6), nullable=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    updated_at = db.Column(db.DateTime, default=db.func.current_timestamp(), onupdate=db.func.current_timestamp())

# A decorator to check if the user is logged in
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not request.cookies.get('logged_in') == 'true':
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Endpoint route to send OTP to valid users
@app.route('/send-otp', methods=['POST'])
def send_otp():
    data = request.json
    email = data.get('email')

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    # Check if the email exists in the database
    user = Users.query.filter_by(email=email).first()
    if not user:
        return jsonify({'error': 'Email not found'}), 404

    otp = generate_otp()

    # Update the OTP in the database
    user.otp = otp
    db.session.commit()

    # Send the OTP email
    try:
        msg = Message('Your OTP Code', sender=os.getenv('MAIL_USERNAME'), recipients=[email])
        msg.body = f'Your OTP code is {otp}.'
        mail.send(msg)
    except Exception as e:
        return jsonify({'error': 'Failed to send OTP. Please try again later.'}), 500

    # Store email in cookies with expiration
    resp = make_response(jsonify({'message': 'OTP sent successfully'}))
    resp.set_cookie('email', email, max_age=600)  # Cookie expires in 10 minutes
    return resp

# Endpoint to check weather a valid otp is entered or not
@app.route('/verify-otp', methods=['POST'])
def verify_otp():
    data = request.json
    email = data.get('email')
    otp = data.get('otp')

    if not email or not otp:
        return jsonify({'error': 'Email and OTP are required'}), 400

    # Check if the email exists
    user = Users.query.filter_by(email=email).first()
    if not user:
        return jsonify({'valid': False}), 400

    # Ensure both OTPs are stripped of extra spaces
    user_otp = user.otp.strip() if user.otp else ''
    otp = otp.strip()

    if user_otp != otp:
        return jsonify({'valid': False}), 400

    # Clear the OTP after successful verification
    user.otp = None
    db.session.commit()

    # Set logged_in cookie
    resp = make_response(jsonify({'valid': True}))
    resp.set_cookie('logged_in', 'true', max_age=1800)  # Cookie expires in 30 minutes

    # Set user_name cookie
    resp.set_cookie('name', user.name, max_age=1800)  # Cookie expires in 30 minutes

    return resp

# Login page route
@app.route('/')
def login():
    return render_template('login.html')

# Route to check OTP
@app.route('/otp')
def otp_page():
    # Check if the user is already logged in
    if request.cookies.get('logged_in') == 'true':
        return redirect(url_for('admin_dashboard'))
    return render_template('otp.html')

# Route for the index page ("Displays invoices")
# Route for the index page("Displays invoice's")
@app.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Initialize the query and filters
    query = "SELECT * FROM invoices"
    filters = {}
    conditions = []

    # Get filter parameter from URL (optional for predefined filters like 'cleared', 'uncleared', etc.)
    filter_type = request.args.get('filter')

    # Process POST data for form filters
    if request.method == 'POST':
        vendor = request.form.get('vendor', '').lower()
        invoice_date = request.form.get('invoice_date')
        date_submission = request.form.get('date_submission')
        invoice_number = request.form.get('invoice_number')
        po_number = request.form.get('po_number')
        created_by = request.form.get('created_by', '').lower()

        # Add conditions for filters
        if vendor:
            conditions.append("vendor LIKE %s")
            filters['vendor'] = f"%{vendor}%"
        if invoice_date:
            conditions.append("invoice_date = %s")
            filters['invoice_date'] = invoice_date
        if date_submission:
            conditions.append("date_submission = %s")
            filters['date_submission'] = date_submission
        if invoice_number:
            conditions.append("invoice_number = %s")
            filters['invoice_number'] = invoice_number
        if po_number:
            conditions.append("po_number = %s")
            filters['po_number'] = po_number    
        if created_by:
            conditions.append("created_by LIKE %s")
            filters['created_by'] = f"%{created_by}%"

    # Apply additional filter based on query parameter
    if filter_type == 'cleared':
        query += " WHERE invoice_cleared = 'Yes'"
    elif filter_type == 'uncleared':
        query += " WHERE invoice_cleared = 'No'"
    elif filter_type == 'all':  # Reset filters to show all invoices
        query = "SELECT * FROM invoices"

    # Add conditions to the query if filters are set
    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # Execute the SQL query
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Check if no invoices are returned
    no_results_message = None
    if not invoices:
        no_results_message = "No matching records found."

    # Fetch values for filter dropdowns (e.g., distinct vendors, etc.)
    cursor.execute("SELECT DISTINCT vendor FROM invoices")
    vendor_values = cursor.fetchall()

    cursor.execute("SELECT DISTINCT created_by FROM invoices")
    created_by_values = cursor.fetchall()

    cursor.execute("SELECT DISTINCT invoice_date FROM invoices")
    invoice_date_values = cursor.fetchall()

    # Close connection
    conn.close()

    # Render the template with necessary context
    return render_template(
        'index.html', 
        invoices=invoices, 
        filters=filters, 
        no_results_message=no_results_message,
        vendor_values=vendor_values,
        created_by_values=created_by_values,
        invoice_date_values=invoice_date_values
    )
    pass

# Route for admin dashboard
@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
def admin_dashboard():
    filters = {}
    query = "SELECT * FROM invoices"

    # Get filter parameter from URL
    filter_type = request.args.get('filter')

    # Add filters based on form input
    if request.method == 'POST':
        vendor = request.form.get('vendor', '').lower()
        invoice_date = request.form.get('invoice_date')
        date_submission = request.form.get('date_submission')
        invoice_number = request.form.get('invoice_number')
        po_number = request.form.get('po_number')
        created_by = request.form.get('created_by', '').lower()

        # New fields for filtering based on HOD approval, CEO approval, and Reviewed By
        hod_approval = request.form.get('hod_approval')
        ceo_approval = request.form.get('ceo_approval')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')

        conditions = []

        if vendor:
            conditions.append("vendor LIKE %s")
            filters['vendor'] = f"%{vendor}%"
        if invoice_date:
            conditions.append("invoice_date = %s")
            filters['invoice_date'] = invoice_date
        if date_submission:
            conditions.append("date_submission = %s")
            filters['date_submission'] = date_submission
        if invoice_number:
            conditions.append("invoice_number = %s")
            filters['invoice_number'] = invoice_number
        if po_number:
            conditions.append("po_number = %s")
            filters['po_number'] = po_number    
        if created_by:
            conditions.append("created_by LIKE %s")
            filters['created_by'] = f"%{created_by}%"

        # Add conditions for HOD approval, CEO approval, and Reviewed By filters
        if hod_approval:
            conditions.append("hod_values = %s")
            filters['hod_approval'] = hod_approval
        if ceo_approval:
            conditions.append("ceo_values = %s")
            filters['ceo_approval'] = ceo_approval
        if reviewed_by:
            conditions.append("reviewed_by = %s")
            filters['reviewed_by'] = reviewed_by

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

    # Apply filters based on the URL query parameter
    if filter_type == 'cleared':
        query += " WHERE invoice_cleared = 'Yes'"
    elif filter_type == 'uncleared':
        query += " WHERE invoice_cleared = 'No'"
    elif filter_type == 'all':  # Reset filters to show all invoices
        query = "SELECT * FROM invoices"

    # Execute the SQL query
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Check if no invoices are returned for the applied filters
    no_results_message = None
    if not invoices:
        no_results_message = "No matching records found."

    # Total number of invoices
    cursor.execute('SELECT COUNT(*) FROM invoices')
    total_invoices = cursor.fetchone()['COUNT(*)']

    # Total cleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "Yes"')
    total_cleared_invoices = cursor.fetchone()['COUNT(*)']

    # Total uncleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "No"')
    total_uncleared_invoices = cursor.fetchone()['COUNT(*)']

    # Calculate Overall Pool (sum of total_amount where invoice_cleared = 'Yes')
    cursor.execute('SELECT SUM(total_amount) as overall_pool FROM invoices WHERE invoice_cleared = "Yes"')
    overall_pool = cursor.fetchone()['overall_pool'] or 0

    # Calculate Monthly Pool (sum of total_amount for the current month where invoice_cleared = 'Yes')
    today = date.today()  # Use date.today() to get current date
    current_month = today.month
    current_year = today.year

    cursor.execute('''
        SELECT SUM(total_amount) as monthly_pool FROM invoices
        WHERE MONTH(invoice_date) = %s AND YEAR(invoice_date) = %s AND invoice_cleared = "Yes"
    ''', (current_month, current_year))
    monthly_pool = cursor.fetchone()['monthly_pool'] or 0

    conn.close()

    # Pass the necessary data to the template
    return render_template(
        'admin_dashboard.html',
        invoices=invoices,
        total_invoices=total_invoices,
        total_cleared_invoices=total_cleared_invoices,
        total_uncleared_invoices=total_uncleared_invoices,
        overall_pool=overall_pool,
        monthly_pool=monthly_pool,
        today=today,  # Pass 'today' to the template
        no_results_message=no_results_message  # Pass message to template
    )
    pass

# Route for download excel button("With / Without filters")
@app.route('/download_excel', methods=['POST'])
def download_excel():
    vendor = request.form.get('vendor')
    invoice_date = request.form.get('invoice_date')
    date_submission = request.form.get('date_submission')
    invoice_number = request.form.get('invoice_number')
    po_number = request.form.get('po_number')
    created_by = request.form.get('created_by')

    # Building the raw SQL query
    query = "SELECT * FROM invoices"
    conditions = []
    filters = {}

    # Apply filters based on form input
    if vendor:
        conditions.append("vendor = %s")
        filters['vendor'] = vendor
    if invoice_date:
        conditions.append("invoice_date = %s")
        filters['invoice_date'] = invoice_date
    if date_submission:
        conditions.append("date_submission = %s")
        filters['date_submission'] = date_submission
    if invoice_number:
        conditions.append("invoice_number = %s")
        filters['invoice_number'] = invoice_number
    if po_number:
        conditions.append("po_number = %s")
        filters['po_number'] = po_number
    if created_by:
        conditions.append("created_by = %s")
        filters['created_by'] = created_by

    # Append conditions to the query
    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # Execute the SQL query
    conn = get_db_connection()  # Replace this with your actual connection function
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Convert to DataFrame for easy export to Excel
    df = pd.DataFrame([{
        'Invoice Date': invoice['invoice_date'],
        'Date Received': invoice['date_received'],
        'Vendor': invoice['vendor'],
        'Invoice Number': invoice['invoice_number'],
        'PO Number': invoice['po_number'],
        'MSME': invoice['msme'],
        'Invoice Amount': invoice['invoice_amount'],
        'GST': invoice['gst'],
        'Total Amount': invoice['total_amount'],
        'Date of Submission': invoice['date_submission'],
        'Approved By': invoice['approved_by'],
        'HOD Approval': invoice['hod_values'],
        'CEO Approval': invoice['ceo_values'],
        'Reviewed By': invoice['reviewed_by'],
        'Created By': invoice['created_by'],
        'Tag1': invoice['tag1'],
        'Tag2': invoice['tag2'],
        'Invoice Cleared': invoice['invoice_cleared'],
        'Cleared Date': invoice['invoice_cleared_date'],
    } for invoice in invoices])

    # Create a bytes buffer for the Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Seek to the beginning of the stream
    output.seek(0)

    # Send the Excel file to the user for download
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='filtered_invoices.xlsx')


    # Execute the SQL query
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Check if no invoices are returned for the applied filters
    no_results_message = None
    if not invoices:
        no_results_message = "No matching records found."

    # Total number of invoices
    cursor.execute('SELECT COUNT(*) FROM invoices')
    total_invoices = cursor.fetchone()['COUNT(*)']

    # Total cleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "Yes"')
    total_cleared_invoices = cursor.fetchone()['COUNT(*)']

    # Total uncleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "No"')
    total_uncleared_invoices = cursor.fetchone()['COUNT(*)']

    conn.close()

    return render_template(
        'admin_dashboard.html', invoices=invoices,
        total_invoices=total_invoices,
        total_cleared_invoices=total_cleared_invoices,
        total_uncleared_invoices=total_uncleared_invoices,
        no_results_message=no_results_message  # Pass message to template
    )
    pass

@app.route('/add', methods=('GET', 'POST'))
@login_required
def add_invoice():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch vendors for the dropdown
    cursor.execute('SELECT * FROM vendors')
    vendors = cursor.fetchall()

    # Fetch approved_by, created_by, hod_values, ceo_values, reviewed_by from .env
    approved_by_values = os.getenv("APPROVED_BY_VALUES", "").split(",")
    created_by_values = os.getenv("CREATED_BY_VALUES", "").split(",")
    hod_values = os.getenv("HOD_VALUES", "").split(",")
    ceo_values = os.getenv("CEO_VALUES", "").split(",")
    reviewed_by_values = os.getenv("REVIEWED_BY_VALUES", "").split(",")
    tag1 = os.getenv("TAG1_VALUES", "").split(",")
    tag2 = os.getenv("TAG2_VALUES", "").split(",")

    if request.method == 'POST':
        # Fetch form data
        invoice_date = request.form['invoice_date']
        date_received = request.form['date_received']
        vendor = request.form['vendor']
        mobile_no = request.form['mobile_no']
        invoice_number = request.form['invoice_number']
        date_submission = request.form['date_submission']
        approved_by = request.form.get('approved_by')
        created_by = request.form['created_by']
        po_approved = request.form['po_approved']
        po_number = request.form['po_number']
        agreement_signed = request.form['agreement_signed']
        po_expiry_date = request.form.get('po_expiry_date') or None
        agreement_signed_date = request.form.get('agreement_signed_date') or None
        hod_approval = request.form.get('hod_values')
        ceo_approval = request.form.get('ceo_values')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')
        invoice_amount = float(request.form['invoice_amount'])
        gst = invoice_amount * 0.18
        total_amount = invoice_amount + gst
        total_amount = round(total_amount, 2)
        
        msme = request.form['msme']

        total_amount_words = num2words(total_amount, to='currency', currency='INR', lang='en_IN').upper().replace(",", "")

        # Validate form inputs
        if not invoice_date or not date_received or not vendor or not invoice_number or not date_submission or not created_by or not invoice_amount:
            flash('All fields are required!')
        else:
            # Check if the invoice number already exists
            cursor.execute('SELECT * FROM invoices WHERE invoice_number = %s', (invoice_number,))
            existing_invoice = cursor.fetchone()

            if existing_invoice:
                flash('Invoice number already exists. Please enter a unique invoice number.')
            else:
                cursor.execute(
                    'INSERT INTO invoices (invoice_date, date_received, vendor, mobile_no, invoice_number, po_approved, po_number, po_expiry_date, agreement_signed, agreement_signed_date, date_submission, approved_by, created_by, tag1, tag2, invoice_amount, gst, total_amount, msme, hod_values, ceo_values, reviewed_by) '
                    'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                    (invoice_date, date_received, vendor, mobile_no, invoice_number, po_approved, po_number, po_expiry_date, agreement_signed, agreement_signed_date, date_submission, approved_by, created_by, tag1, tag2, invoice_amount, gst, total_amount, msme, hod_approval, ceo_approval, reviewed_by)
                )
                conn.commit()
                flash('Invoice added successfully!')

                # Generate Excel file using openpyxl
                template_path = "static/excel_templates/template.xlsx"

                wb = load_workbook(template_path)
                ws = wb.active

                # Populate data into the template
                ws['A7'] = vendor
                ws['E6'] = invoice_date
                ws['E7'] = invoice_number
                ws['B8'] = po_approved
                ws['B9'] = agreement_signed
                ws['E8'] = po_expiry_date
                ws['E9'] = agreement_signed_date
                ws['E11'] = msme
                ws['E12'] = mobile_no
                ws['F25'] = total_amount
                ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F25'].number_format = numbers.FORMAT_NUMBER_00

                ws['F35'] = total_amount
                ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F35'].number_format = numbers.FORMAT_NUMBER_00

                # ws['B35'] = total_amount_words
                if ws['A35'].value:
                    ws['A35'] = f"{ws['A35'].value} {total_amount_words}"
                else:
                    ws['A35'] = total_amount_words
                ws['A35'].alignment = Alignment(wrap_text=True)
                ws.column_dimensions['A'].width = 50 # Increase the width of column A
                ws.row_dimensions[35].height = 40  # Increase the height of row 35

                if ws['A25'].value:  # Check if there is already a value in the cell
                    ws['A25'] = f"{ws['A25'].value} ({tag1})"
                else:
                    ws['A25'] = f"Marketing Expenses ({tag1})"

                # ws['D11'] = date_submission

                if ws['D38'].value:  # Check if there's already a value in the cell
                    ws['D38'] = f"{ws['D38'].value} {created_by}"
                else:
                    ws['D38'] = created_by

                # For approved_by
                if ws['D40'].value:
                    ws['D40'] = f"{ws['D40'].value} {approved_by}"
                else:
                    ws['D40'] = approved_by

                # For reviewed_by
                if ws['D39'].value:
                    ws['D39'] = f"{ws['D39'].value} {reviewed_by}"
                else:
                    ws['D39'] = reviewed_by

                # For hod_approval
                if ws['D42'].value:
                    ws['D42'] = f"{ws['D42'].value} {hod_approval}"
                else:
                    ws['D42'] = hod_approval

                # For ceo_approval
                if ws['D44'].value:
                    ws['D44'] = f"{ws['D44'].value} {ceo_approval}"
                else:
                    ws['D44'] = ceo_approval

                # Save the populated Excel file to a BytesIO stream
                excel_file = BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                # Return the Excel file as a response
                return send_file(
                    excel_file,
                    as_attachment=True,
                    download_name=f"invoice_{invoice_number}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    conn.close()

    return render_template(
        'add_invoice.html',
        vendors=vendors,
        approved_by_values=approved_by_values,
        created_by_values=created_by_values,
        hod_values=hod_values,
        ceo_values=ceo_values,
        reviewed_by_values=reviewed_by_values,
        tag1=tag1,
        tag2=tag2
    )
    pass

@app.route('/edit/<int:id>', methods=('GET', 'POST'))
@login_required
def edit_invoice(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch the invoice to be edited
    cursor.execute('SELECT * FROM invoices WHERE id = %s', (id,))
    invoice = cursor.fetchone()

    # Fetch vendors for the dropdown
    cursor.execute('SELECT * FROM vendors')
    vendors = cursor.fetchall()

    # Fetch approved_by and created_by values from .env
    approved_by_values = os.getenv("APPROVED_BY_VALUES", "").split(",")
    created_by_values = os.getenv("CREATED_BY_VALUES", "").split(",")
    hod_values = os.getenv("HOD_VALUES", "").split(",")
    ceo_values = os.getenv("CEO_VALUES", "").split(",")
    reviewed_by_values = os.getenv("REVIEWED_BY_VALUES", "").split(",")
    tag1 = os.getenv("TAG1_VALUES", "").split(",")
    tag2 = os.getenv("TAG2_VALUES", "").split(",")

    if request.method == 'POST':
        invoice_date = request.form['invoice_date']
        date_received = request.form['date_received']
        vendor = request.form['vendor']
        mobile_no = request.form['mobile_no']
        invoice_number = request.form['invoice_number']
        po_approved = request.form['po_approved']
        po_number = request.form['po_number']
        agreement_signed = request.form['agreement_signed']
        date_submission = request.form['date_submission']
        approved_by = request.form.get('approved_by')
        hod_approval = request.form.get('hod_values')
        ceo_approval = request.form.get('ceo_values')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')
        po_expiry_date = request.form.get('po_expiry_date')  # Optional
        agreement_signed_date = request.form.get('agreement_signed_date')  # Optional
        po_expiry_date = po_expiry_date if po_expiry_date else None
        agreement_signed_date = agreement_signed_date if agreement_signed_date else None
        created_by = invoice['created_by']
        invoice_cleared = request.form['invoice_cleared']

        if invoice_cleared == 'Yes':
            invoice_cleared_date = request.form.get('invoice_cleared_date') or date.today()
        else:
            invoice_cleared_date = None

        invoice_amount = float(request.form['invoice_amount'])
        gst = invoice_amount * 0.18
        total_amount = invoice_amount + gst
        total_amount_words = num2words(total_amount, to='currency', currency='INR', lang='en_IN')
        total_amount_words = total_amount_words.upper().replace(",", "")
        msme = request.form['msme']

        if not invoice_date or not date_received or not vendor or not invoice_number or not date_submission or not invoice_amount:
            flash('All fields are required!')
        else:
            cursor.execute(
                '''
                UPDATE invoices 
                SET invoice_date = %s, 
                    date_received = %s, 
                    vendor = %s, 
                    mobile_no = %s,
                    invoice_number = %s, 
                    po_approved = %s, 
                    po_number = %s, 
                    po_expiry_date = %s, 
                    agreement_signed = %s, 
                    agreement_signed_date = %s, 
                    date_submission = %s, 
                    approved_by = %s, 
                    created_by = %s, 
                    tag1 = %s, 
                    tag2 = %s, 
                    invoice_amount = %s, 
                    gst = %s, 
                    total_amount = %s, 
                    msme = %s, 
                    hod_values = %s, 
                    ceo_values = %s, 
                    reviewed_by = %s, 
                    invoice_cleared = %s, 
                    invoice_cleared_date = %s
                WHERE id = %s
                ''',
                (
                    invoice_date, date_received, vendor, mobile_no, invoice_number, 
                    po_approved, po_number, po_expiry_date, agreement_signed, 
                    agreement_signed_date, date_submission, approved_by, created_by, 
                    tag1, tag2, invoice_amount, gst, total_amount, msme, 
                    hod_approval, ceo_approval, reviewed_by, invoice_cleared, 
                    invoice_cleared_date, id
                )
            )
            conn.commit()

            flash('Invoice updated successfully!')

            # Skip Excel generation if invoice is cleared
            if invoice_cleared == 'Yes':
                conn.close()
                return redirect(url_for('index'))  # Redirect to a different view after successful update
            
            # Generate the updated invoice Excel
            template_path = "static/excel_templates/template.xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            # Populate data into the template
            ws['A7'] = vendor
            ws['E6'] = invoice_date
            ws['E7'] = invoice_number
            ws['B8'] = po_approved
            ws['B9'] = agreement_signed
            ws['E8'] = po_expiry_date
            ws['E9'] = agreement_signed_date
            ws['E11'] = msme
            ws['E12'] = mobile_no
            ws['F25'] = total_amount
            ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F25'].number_format = numbers.FORMAT_NUMBER_00

            ws['F35'] = total_amount
            ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F35'].number_format = numbers.FORMAT_NUMBER_00

            if ws['A35'].value:
                ws['A35'] = f"{ws['A35'].value} {total_amount_words}"
            else:
                ws['A35'] = total_amount_words
            ws['A35'].alignment = Alignment(wrap_text=True)
            ws.column_dimensions['A'].width = 50  # Increase the width of column A
            ws.row_dimensions[35].height = 40  # Increase the height of row 35

            if ws['A25'].value:  # Check if there is already a value in the cell
                ws['A25'] = f"{ws['A25'].value} ({tag1})"
            else:
                ws['A25'] = f"Marketing Expenses ({tag1})"

            if ws['D38'].value:  # Check if there's already a value in the cell
                ws['D38'] = f"{ws['D38'].value} {created_by}"
            else:
                ws['D38'] = created_by

            # For approved_by
            if ws['D40'].value:
                ws['D40'] = f"{ws['D40'].value} {approved_by}"
            else:
                ws['D40'] = approved_by

            # For reviewed_by
            if ws['D39'].value:
                ws['D39'] = f"{ws['D39'].value} {reviewed_by}"
            else:
                ws['D39'] = reviewed_by

            # For hod_approval
            if ws['D42'].value:
                ws['D42'] = f"{ws['D42'].value} {hod_approval}"
            else:
                ws['D42'] = hod_approval

            # For ceo_approval
            if ws['D44'].value:
                ws['D44'] = f"{ws['D44'].value} {ceo_approval}"
            else:
                ws['D44'] = ceo_approval

            # Save the populated Excel file to a BytesIO stream
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Return the Excel file as a response
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=f"updated_invoice_{invoice_number}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    conn.close()

    return render_template(
        'edit_invoice.html', 
        invoice=invoice, 
        vendors=vendors, 
        approved_by_values=approved_by_values, 
        created_by_values=created_by_values,
        hod_values=hod_values,
        ceo_values=ceo_values,
        reviewed_by_values=reviewed_by_values,
        tag1=tag1,
        tag2=tag2
    )
    pass

# Route to delete invoices
@app.route('/delete/<int:id>', methods=('POST',))
@login_required
def delete_invoice(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM invoices WHERE id = %s', (id,))
    conn.commit()
    conn.close()
    flash('Invoice deleted successfully!')
    return redirect(url_for('index'))
    pass
    

# Route to manage the vendors("Add vendors")
@app.route('/manage_vendors', methods=['GET', 'POST'])
@login_required
def manage_vendors():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        new_vendor = request.form['new_vendor']
        vendor_status = request.form['vendor_status']  # Get the vendor status from the form

        if new_vendor:
            cursor.execute('INSERT INTO vendors (vendor_name, vendor_status) VALUES (%s, %s)', 
                           (new_vendor, vendor_status))  # Insert vendor status along with name
            conn.commit()
            flash('Vendor added successfully!')
        else:
            flash('Vendor name cannot be empty!')

    # Fetch all vendors along with their status
    cursor.execute('SELECT * FROM vendors')
    vendors = cursor.fetchall()
    conn.close()

    return render_template('manage_vendors.html', vendors=vendors)
    pass

# Route to Edit the vendors
@app.route('/edit_vendor/<int:id>', methods=['POST'])
@login_required
def edit_vendor(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    vendor_status = request.form['vendor_status']
    
    cursor.execute('UPDATE vendors SET vendor_status = %s WHERE id = %s', (vendor_status, id))
    conn.commit()
    conn.close()

    flash('Vendor status updated successfully!')
    return redirect(url_for('manage_vendors'))
    pass

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
